"""
=============================================================
  DESCARGA MASIVA DE CFDIs - SAT México  v3.0
  Facturas EMITIDAS y RECIBIDAS + Rangos inteligentes
=============================================================

  LÓGICA DE RANGOS AUTOMÁTICA:
  ┌─────────────────────────────────────────────────────────┐
  │ 1ª ejecución  → descarga desde 01-Ene del año actual   │
  │                 hasta hoy                               │
  │                                                         │
  │ Ejecuciones   → desde (última fecha descargada - 1 día) │
  │ siguientes      hasta hoy  (asegura sin huecos)         │
  │                                                         │
  │ Si un CFDI ya existe (mismo UUID) → overwrite           │
  └─────────────────────────────────────────────────────────┘

  USO:
    python descarga_cfdi_sat.py                          # rango automático
    python descarga_cfdi_sat.py --inicio 2025-01-01      # desde fecha fija
    python descarga_cfdi_sat.py --inicio 2025-01-01 --fin 2025-03-31
    python descarga_cfdi_sat.py --auto                   # scheduler mensual

  AUTOMATIZACIÓN WINDOWS:
    Ejecuta configurar_tarea_windows.bat como Administrador

  AUTOMATIZACIÓN MAC/LINUX:
    crontab -e
    0 8 1 * * python3 /ruta/descarga_cfdi_sat.py

  ADVERTENCIA: Nunca compartas tus archivos .cer, .key ni contraseña.
=============================================================
"""

import argparse
import base64
import datetime
import json
import logging
import os
import sys
import time
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path


# ══════════════════════════════════════════════════════════════
#  DEPENDENCIAS AUTOMÁTICAS
# ══════════════════════════════════════════════════════════════

def _instalar(pkg):
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

for _pkg in ["cfdiclient", "openpyxl", "lxml", "schedule"]:
    try:
        __import__(_pkg.replace("-", "_"))
    except ImportError:
        print(f"📦 Instalando {_pkg}...")
        _instalar(_pkg)

from cfdiclient import (
    Autenticacion, DescargaMasiva, Fiel,
    SolicitaDescarga, VerificaSolicitudDescarga,
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import schedule


# ══════════════════════════════════════════════════════════════
#  ✏️  CONFIGURACIÓN — Edita estos valores
# ══════════════════════════════════════════════════════════════

RFC             = "XAXX010101000"          # Tu RFC (sin espacios)
FIEL_CER        = "efirma/tu_archivo.cer"  # Ruta a tu .cer
FIEL_KEY        = "efirma/tu_archivo.key"  # Ruta a tu .key
FIEL_PASSWORD   = "tu_contraseña"          # Contraseña e.firma

CARPETA_DESTINO = Path("contabilidad_sat") # Carpeta raíz de todo
TIPO_SOLICITUD  = "CFDI"                   # "CFDI" o "Metadata"
ESPERA_SEGUNDOS = 90                       # Segundos entre reintentos SAT

DIA_AUTO        = 1                        # Día del mes para descarga auto
HORA_AUTO       = "08:00"                  # Hora de descarga automática

# ══════════════════════════════════════════════════════════════


# ── Logging ────────────────────────────────────────────────────
CARPETA_DESTINO.mkdir(parents=True, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(CARPETA_DESTINO / "descarga_sat.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════
#  HISTORIAL  —  fuente de verdad de rangos y UUIDs descargados
# ══════════════════════════════════════════════════════════════

HISTORIAL_FILE = CARPETA_DESTINO / "historial_descargas.json"

HISTORIAL_SCHEMA = {
    "ultima_fecha_descargada": None,   # "YYYY-MM-DD" última fecha_fin procesada
    "primera_ejecucion": True,         # True hasta que se complete 1ª descarga
    "uuids_descargados": [],           # Lista de UUIDs ya guardados en disco
    "ejecuciones": [],                 # Log de cada ejecución [{fecha, inicio, fin, total}]
}


def cargar_historial() -> dict:
    if HISTORIAL_FILE.exists():
        datos = json.loads(HISTORIAL_FILE.read_text(encoding="utf-8"))
        # Migración: asegurar claves nuevas si viene de versión anterior
        for k, v in HISTORIAL_SCHEMA.items():
            datos.setdefault(k, v)
        return datos
    return dict(HISTORIAL_SCHEMA)


def guardar_historial(h: dict):
    HISTORIAL_FILE.write_text(
        json.dumps(h, indent=2, ensure_ascii=False, default=str),
        encoding="utf-8",
    )


# ══════════════════════════════════════════════════════════════
#  LÓGICA DE RANGO INTELIGENTE
# ══════════════════════════════════════════════════════════════

def calcular_rango(
    inicio_manual: datetime.date = None,
    fin_manual:    datetime.date = None,
) -> tuple[datetime.date, datetime.date]:
    """
    Determina el rango de descarga según las reglas:

    - Si se pasan --inicio / --fin  → se usan tal cual (modo manual)
    - Si es la primera ejecución    → 01-Ene-año_actual → hoy
    - Si ya hay descargas previas   → (última_fecha - 1 día) → hoy
                                      (el -1 día evita huecos por zona horaria)
    """
    hoy = datetime.date.today()

    if inicio_manual and fin_manual:
        log.info("📅 Rango MANUAL proporcionado por parámetros.")
        return inicio_manual, fin_manual

    if inicio_manual and not fin_manual:
        log.info("📅 Rango MANUAL: inicio fijo, fin = hoy.")
        return inicio_manual, hoy

    h = cargar_historial()

    if h["primera_ejecucion"] or not h["ultima_fecha_descargada"]:
        inicio = datetime.date(hoy.year, 1, 1)
        log.info(f"📅 Primera ejecución — rango: {inicio} → {hoy}")
        return inicio, hoy

    ultima = datetime.date.fromisoformat(h["ultima_fecha_descargada"])
    inicio = ultima - datetime.timedelta(days=1)
    log.info(f"📅 Ejecución incremental — rango: {inicio} → {hoy}  "
             f"(última descarga: {ultima})")
    return inicio, hoy


# ══════════════════════════════════════════════════════════════
#  OVERWRITE INTELIGENTE POR UUID
# ══════════════════════════════════════════════════════════════

def uuid_desde_xml(ruta: Path) -> str | None:
    """Extrae el UUID del timbre fiscal de un XML sin parsearlo completo."""
    try:
        ns = {"tfd": "http://www.sat.gob.mx/TimbreFiscalDigital"}
        root = ET.parse(ruta).getroot()
        tfd = root.find(".//tfd:TimbreFiscalDigital", ns)
        return tfd.get("UUID") if tfd is not None else None
    except Exception:
        return None


def guardar_xml_con_overwrite(
    contenido: bytes,
    carpeta:   Path,
    nombre:    str,
    uuid:      str,
    uuids_set: set,
) -> tuple[bool, str]:
    """
    Guarda un XML en disco.
    - Si el UUID ya existe: sobreescribe (overwrite) y registra en log.
    - Si es nuevo: lo crea y agrega el UUID al set.
    Retorna (es_nuevo, ruta_guardada).
    """
    ruta = carpeta / nombre
    es_nuevo = uuid not in uuids_set

    if not es_nuevo:
        log.debug(f"   ↩️  Overwrite: {nombre} (UUID ya existía)")
    else:
        log.debug(f"   ➕ Nuevo:     {nombre}")

    ruta.write_bytes(contenido)
    uuids_set.add(uuid)
    return es_nuevo, str(ruta)


# ══════════════════════════════════════════════════════════════
#  ESTRUCTURA DE CARPETAS
# ══════════════════════════════════════════════════════════════

def crear_estructura(fecha_ini: datetime.date, fecha_fin: datetime.date) -> dict:
    """
    Crea carpetas agrupadas por rango de fechas.
    Si el rango abarca varios meses se usa la carpeta del mes de inicio.
    """
    base = CARPETA_DESTINO / str(fecha_ini.year) / f"{fecha_ini.month:02d}"
    carpetas = {
        "recibidas_xml": base / "recibidas" / "xml",
        "recibidas_zip": base / "recibidas" / "zips",
        "emitidas_xml":  base / "emitidas"  / "xml",
        "emitidas_zip":  base / "emitidas"  / "zips",
        "reportes":      base / "reportes",
    }
    for c in carpetas.values():
        c.mkdir(parents=True, exist_ok=True)
    log.info(f"📁 Estructura: {base.resolve()}")
    return carpetas


# ══════════════════════════════════════════════════════════════
#  AUTENTICACIÓN E.FIRMA
# ══════════════════════════════════════════════════════════════

def cargar_fiel() -> Fiel:
    log.info("🔑 Cargando e.firma...")
    for ruta in [FIEL_CER, FIEL_KEY]:
        if not Path(ruta).exists():
            raise FileNotFoundError(f"Archivo no encontrado: {ruta}")
    fiel = Fiel(
        open(FIEL_CER, "rb").read(),
        open(FIEL_KEY, "rb").read(),
        FIEL_PASSWORD,
    )
    log.info("✅ e.firma OK")
    return fiel


def obtener_token(fiel: Fiel) -> str:
    token = Autenticacion(fiel).obtener_token()
    if not token:
        raise Exception("Token no obtenido. Verifica tu e.firma.")
    return token


# ══════════════════════════════════════════════════════════════
#  WEB SERVICE SAT
# ══════════════════════════════════════════════════════════════

def solicitar(fiel, fecha_ini, fecha_fin, tipo) -> str:
    token = obtener_token(fiel)
    desc  = SolicitaDescarga(fiel)
    kwargs = dict(
        token=token,
        rfc_solicitante=RFC,
        fecha_inicial=fecha_ini,
        fecha_final=fecha_fin,
        tipo_solicitud=TIPO_SOLICITUD,
    )
    if tipo == "recibidas":
        kwargs["rfc_receptor"] = RFC
        log.info(f"📤 Solicitando RECIBIDAS {fecha_ini} → {fecha_fin}")
    else:
        kwargs["rfc_emisor"] = RFC
        log.info(f"📤 Solicitando EMITIDAS  {fecha_ini} → {fecha_fin}")

    resp   = desc.solicitar_descarga(**kwargs)
    id_sol = resp.get("IdSolicitud") or resp.get("id_solicitud")
    codigo = resp.get("CodEstatus")  or resp.get("cod_estatus")
    log.info(f"   ID solicitud: {id_sol}  |  Código SAT: {codigo}")

    if not id_sol:
        raise Exception(f"SAT no devolvió ID: {resp}")
    return id_sol


def verificar(fiel, id_sol) -> list:
    log.info(f"⏳ Verificando {id_sol} ...")
    verif = VerificaSolicitudDescarga(fiel)

    for intento in range(1, 200):
        token    = obtener_token(fiel)
        res      = verif.verificar_descarga(token, RFC, id_sol)
        estado   = str(res.get("EstadoSolicitud", ""))
        codigo   = res.get("CodEstatus", "")
        paquetes = res.get("IdsPaquetes", [])

        log.info(f"   [{intento}] estado={estado} código={codigo} paquetes={len(paquetes)}")

        if   estado == "3":          return paquetes      # Terminada ✅
        elif estado == "5":
            log.warning("   Sin comprobantes en ese período.")
            return []
        elif estado in ["1", "2"]:   time.sleep(ESPERA_SEGUNDOS)
        else:
            raise Exception(f"Error SAT estado={estado} código={codigo}")

    raise Exception("Tiempo de espera agotado sin respuesta del SAT.")


def descargar_paquetes(fiel, paquetes, carpeta_zip) -> list:
    log.info(f"📥 Descargando {len(paquetes)} paquete(s)...")
    desc = DescargaMasiva(fiel)
    zips = []
    for i, pkg_id in enumerate(paquetes, 1):
        token = obtener_token(fiel)
        res   = desc.descargar_paquete(token, RFC, pkg_id)
        b64   = res.get("Paquete") or res.get("paquete")
        if not b64:
            log.warning(f"   Paquete vacío: {pkg_id}")
            continue
        ruta = carpeta_zip / f"{pkg_id}.zip"
        ruta.write_bytes(base64.b64decode(b64))
        zips.append(ruta)
        log.info(f"   ✅ [{i}/{len(paquetes)}] {ruta.name}")
    return zips


def extraer_con_overwrite(
    zips:       list,
    carpeta_xml: Path,
    uuids_set:  set,
) -> tuple[list, int, int]:
    """
    Extrae XMLs de los ZIPs aplicando overwrite por UUID.
    Retorna (lista_rutas, nuevos, sobreescritos).
    """
    log.info(f"📂 Extrayendo XMLs → {carpeta_xml}")
    rutas   = []
    nuevos  = 0
    overwr  = 0

    for ruta_zip in zips:
        with zipfile.ZipFile(ruta_zip) as z:
            nombres_xml = [n for n in z.namelist() if n.lower().endswith(".xml")]
            for nombre in nombres_xml:
                contenido = z.read(nombre)
                # Extraer a temporal para leer UUID
                ruta_temp = carpeta_xml / nombre
                ruta_temp.write_bytes(contenido)
                uuid = uuid_desde_xml(ruta_temp)

                if uuid:
                    es_nuevo, _ = guardar_xml_con_overwrite(
                        contenido, carpeta_xml, nombre, uuid, uuids_set
                    )
                    if es_nuevo:
                        nuevos += 1
                    else:
                        overwr += 1
                else:
                    # Sin UUID (no es CFDI válido), igual se guarda
                    nuevos += 1

                rutas.append(carpeta_xml / nombre)

            log.info(f"   {len(nombres_xml)} XMLs ← {ruta_zip.name}")

    log.info(f"   Nuevos: {nuevos}  |  Overwrite: {overwr}")
    return rutas, nuevos, overwr


# ══════════════════════════════════════════════════════════════
#  PARSEO CFDI
# ══════════════════════════════════════════════════════════════

NS_CFDI = {
    "cfdi":  "http://www.sat.gob.mx/cfd/4",
    "cfdi3": "http://www.sat.gob.mx/cfd/3",
    "tfd":   "http://www.sat.gob.mx/TimbreFiscalDigital",
}


def parsear_cfdi(ruta: Path, tipo_factura: str) -> dict:
    try:
        root = ET.parse(ruta).getroot()
        g    = root.get

        def nodo(tag):
            return (root.find(f"cfdi:{tag}", NS_CFDI)
                    or root.find(f"cfdi3:{tag}", NS_CFDI))

        emisor   = nodo("Emisor")
        receptor = nodo("Receptor")
        concepto = (root.find("cfdi:Conceptos/cfdi:Concepto",   NS_CFDI)
                    or root.find("cfdi3:Conceptos/cfdi3:Concepto", NS_CFDI))
        tfd      = root.find(".//tfd:TimbreFiscalDigital", NS_CFDI)

        def a(el, attr): return el.get(attr, "") if el is not None else ""

        fecha = g("Fecha", "")
        return {
            "Tipo Factura":    tipo_factura,
            "UUID":            a(tfd,      "UUID"),
            "Fecha":           fecha[:10],
            "Año":             fecha[:4],
            "Mes":             fecha[5:7],
            "Serie":           g("Serie",  ""),
            "Folio":           g("Folio",  ""),
            "RFC Emisor":      a(emisor,   "Rfc"),
            "Nombre Emisor":   a(emisor,   "Nombre"),
            "RFC Receptor":    a(receptor, "Rfc"),
            "Nombre Receptor": a(receptor, "Nombre"),
            "Uso CFDI":        a(receptor, "UsoCFDI"),
            "Descripción":     a(concepto, "Descripcion"),
            "SubTotal":        _to_float(g("SubTotal", "")),
            "Descuento":       _to_float(g("Descuento", "")),
            "Total":           _to_float(g("Total", "")),
            "Moneda":          g("Moneda", "MXN"),
            "Tipo Comprobante":g("TipoDeComprobante", ""),
            "Método Pago":     g("MetodoPago", ""),
            "Forma Pago":      g("FormaPago", ""),
            "Archivo":         ruta.name,
        }
    except Exception as e:
        return {"Tipo Factura": tipo_factura, "Archivo": ruta.name, "Error": str(e)}


def _to_float(val: str):
    try:    return float(val) if val else ""
    except: return val


# ══════════════════════════════════════════════════════════════
#  REPORTE EXCEL
# ══════════════════════════════════════════════════════════════

COLS = [
    "Tipo Factura", "UUID", "Fecha", "Año", "Mes", "Serie", "Folio",
    "RFC Emisor", "Nombre Emisor", "RFC Receptor", "Nombre Receptor",
    "Uso CFDI", "Descripción", "SubTotal", "Descuento", "Total",
    "Moneda", "Tipo Comprobante", "Método Pago", "Forma Pago", "Archivo",
]

C_HEADER   = "1F4E79"
C_EMITIDA  = "C6EFCE"
C_RECIBIDA = "DDEBF7"


def _suma(lst): return sum(d.get("Total") or 0 for d in lst if isinstance(d.get("Total"), float))


def generar_excel(datos: list, carpeta: Path, fecha_ini: datetime.date, fecha_fin: datetime.date):
    log.info("📊 Generando reporte Excel...")
    wb  = openpyxl.Workbook()
    hdr = Font(color="FFFFFF", bold=True)
    fill_h = PatternFill("solid", fgColor=C_HEADER)

    emitidas  = [d for d in datos if d.get("Tipo Factura") == "Emitida"  and "Error" not in d]
    recibidas = [d for d in datos if d.get("Tipo Factura") == "Recibida" and "Error" not in d]

    # ── Hoja Resumen ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    resumen = [
        ("Período",              f"{fecha_ini}  →  {fecha_fin}"),
        ("RFC",                  RFC),
        ("Fecha de generación",  datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Facturas emitidas",    len(emitidas)),
        ("Total emitido (MXN)",  _suma(emitidas)),
        ("", ""),
        ("Facturas recibidas",   len(recibidas)),
        ("Total recibido (MXN)", _suma(recibidas)),
        ("", ""),
        ("Balance (emit-recib)", _suma(emitidas) - _suma(recibidas)),
    ]
    for r, (etiq, val) in enumerate(resumen, 1):
        ws.cell(r, 1, etiq).font = Font(bold=bool(etiq))
        ws.cell(r, 2, val)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30

    # ── Función hoja detalle ──────────────────────────────────
    def hoja_detalle(nombre, filas, color_par):
        wsd = wb.create_sheet(nombre)
        for c, titulo in enumerate(COLS, 1):
            cell = wsd.cell(1, c, titulo)
            cell.fill = fill_h
            cell.font = hdr
            cell.alignment = Alignment(horizontal="center")
        fill_par = PatternFill("solid", fgColor=color_par)
        for r, d in enumerate(filas, 2):
            for c, col in enumerate(COLS, 1):
                wsd.cell(r, c, d.get(col, ""))
            if r % 2 == 0:
                for c in range(1, len(COLS) + 1):
                    wsd.cell(r, c).fill = fill_par
        for col in wsd.columns:
            ancho = max((len(str(c.value or "")) for c in col), default=10)
            wsd.column_dimensions[col[0].column_letter].width = min(ancho + 3, 55)
        wsd.auto_filter.ref = wsd.dimensions
        log.info(f"   Hoja '{nombre}': {len(filas)} registros")

    hoja_detalle("Emitidas",  emitidas,  C_EMITIDA)
    hoja_detalle("Recibidas", recibidas, C_RECIBIDA)

    nombre = f"CFDIs_{fecha_ini}_{fecha_fin}_{RFC}.xlsx"
    ruta   = carpeta / nombre
    wb.save(ruta)
    log.info(f"✅ Excel: {ruta}")
    return ruta


# ══════════════════════════════════════════════════════════════
#  FLUJO PRINCIPAL
# ══════════════════════════════════════════════════════════════

def ejecutar_descarga(
    inicio_manual: datetime.date = None,
    fin_manual:    datetime.date = None,
):
    hoy = datetime.date.today()
    fecha_ini, fecha_fin = calcular_rango(inicio_manual, fin_manual)

    log.info("=" * 62)
    log.info(f"  🧾 DESCARGA CFDI SAT v3.0  |  {fecha_ini} → {fecha_fin}")
    log.info("=" * 62)

    historial  = cargar_historial()
    uuids_set  = set(historial.get("uuids_descargados", []))
    carpetas   = crear_estructura(fecha_ini, fecha_fin)
    fiel       = cargar_fiel()
    todos      = []
    tot_nuevos = 0
    tot_overwr = 0

    for tipo, carp_xml, carp_zip in [
        ("recibidas", carpetas["recibidas_xml"], carpetas["recibidas_zip"]),
        ("emitidas",  carpetas["emitidas_xml"],  carpetas["emitidas_zip"]),
    ]:
        log.info(f"\n{'─' * 50}")
        log.info(f"  ▶  {tipo.upper()}")
        log.info(f"{'─' * 50}")
        try:
            id_sol   = solicitar(fiel, fecha_ini, fecha_fin, tipo)
            paquetes = verificar(fiel, id_sol)
            if not paquetes:
                log.info(f"  Sin paquetes para {tipo}.")
                continue

            zips = descargar_paquetes(fiel, paquetes, carp_zip)
            xmls, nuevos, overwr = extraer_con_overwrite(zips, carp_xml, uuids_set)

            tipo_label = "Emitida" if tipo == "emitidas" else "Recibida"
            datos = [parsear_cfdi(x, tipo_label) for x in xmls]
            todos += datos
            tot_nuevos += nuevos
            tot_overwr += overwr

            log.info(f"  ✅ {tipo.capitalize()}: {len(datos)} CFDIs "
                     f"({nuevos} nuevos, {overwr} overwrite)")

        except Exception as e:
            log.error(f"  ❌ Error en {tipo}: {e}")

    # ── Reporte Excel ─────────────────────────────────────────
    if todos:
        generar_excel(todos, carpetas["reportes"], fecha_ini, fecha_fin)

    # ── Actualizar historial ──────────────────────────────────
    historial["primera_ejecucion"]      = False
    historial["ultima_fecha_descargada"] = str(fecha_fin)
    historial["uuids_descargados"]       = list(uuids_set)
    historial["ejecuciones"].append({
        "fecha_ejecucion": datetime.datetime.now().isoformat(),
        "periodo_inicio":  str(fecha_ini),
        "periodo_fin":     str(fecha_fin),
        "total_cfdi":      len(todos),
        "nuevos":          tot_nuevos,
        "overwrite":       tot_overwr,
    })
    guardar_historial(historial)

    log.info("\n" + "=" * 62)
    log.info(f"  ✅ COMPLETADO")
    log.info(f"     CFDIs procesados : {len(todos)}")
    log.info(f"     Nuevos           : {tot_nuevos}")
    log.info(f"     Overwrite        : {tot_overwr}")
    log.info(f"     Carpeta          : {CARPETA_DESTINO.resolve()}")
    log.info(f"     Próxima descarga : desde {fecha_fin - datetime.timedelta(days=1)}")
    log.info("=" * 62)


# ══════════════════════════════════════════════════════════════
#  SCHEDULER MENSUAL
# ══════════════════════════════════════════════════════════════

def tarea_programada():
    if datetime.date.today().day == DIA_AUTO:
        log.info("🕐 Ejecutando descarga programada mensual...")
        try:
            ejecutar_descarga()
        except Exception as e:
            log.error(f"❌ Error en tarea programada: {e}")


def iniciar_scheduler():
    log.info(f"⏰ Scheduler activo — descarga el día {DIA_AUTO} "
             f"de cada mes a las {HORA_AUTO}")
    log.info("   Deja este proceso corriendo. Ctrl+C para detener.")
    schedule.every().day.at(HORA_AUTO).do(tarea_programada)
    while True:
        schedule.run_pending()
        time.sleep(60)


# ══════════════════════════════════════════════════════════════
#  ENTRADA
# ══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    p = argparse.ArgumentParser(
        description="Descarga masiva de CFDIs del SAT v3.0 — rangos inteligentes"
    )
    p.add_argument("--auto",    action="store_true",
                   help="Modo scheduler mensual en segundo plano")
    p.add_argument("--inicio",  type=str, default=None,
                   help="Fecha inicio YYYY-MM-DD  (omitir = automático)")
    p.add_argument("--fin",     type=str, default=None,
                   help="Fecha fin   YYYY-MM-DD  (omitir = hoy)")
    args = p.parse_args()

    if args.auto:
        iniciar_scheduler()
    else:
        fi = datetime.date.fromisoformat(args.inicio) if args.inicio else None
        ff = datetime.date.fromisoformat(args.fin)    if args.fin    else None
        ejecutar_descarga(fi, ff)
