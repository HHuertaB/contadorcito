"""
ContaSAT - app.py v3.0
Backend Flask. Sirve la GUI en localhost y expone /api/* endpoints.
Sin pywebview, sin pythonnet, sin compilar nada.
Ejecutar: python app.py
"""

import base64
import datetime
import json
import logging
import os
import sys
import threading
import webbrowser
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

# ── Verificar dependencias ────────────────────────────────────
def _verificar_deps():
    faltantes = []
    for m in ["flask", "satcfdi", "openpyxl"]:
        try:
            __import__(m)
        except ImportError:
            faltantes.append(m)
    if faltantes:
        print(f"[ERROR] Faltan dependencias: {faltantes}")
        print("Ejecuta iniciar_contasat.bat para instalarlas.")
        input("Presiona Enter para cerrar...")
        sys.exit(1)

_verificar_deps()

from flask import Flask, jsonify, request, send_file, send_from_directory
from satcfdi.models import Signer
from satcfdi.pacs.sat import SAT, TipoDescargaMasivaTerceros, EstadoSolicitud
import openpyxl
from openpyxl.styles import Font, PatternFill

# ── Rutas ─────────────────────────────────────────────────────
BASE_DIR  = Path(__file__).parent
GUI_FILE  = BASE_DIR / "contasat_gui.html"
DATA_DIR  = BASE_DIR.parent / "contabilidad_sat"
HIST_FILE = DATA_DIR / "historial.json"
LOG_FILE  = DATA_DIR / "descarga_sat.log"
CFG_FILE  = BASE_DIR.parent / "config.json"
PORT      = 5120

DATA_DIR.mkdir(parents=True, exist_ok=True)

# ── Logging ───────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("contasat")

# ── Estado de sesion ─────────────────────────────────────────
_signer      = None
_sat         = None
_log_lines   = []
_progreso    = 0
_descargando = False

# ── Archivo de solicitudes pendientes ────────────────────────
PENDING_FILE = DATA_DIR / "solicitudes_pendientes.json"

def _load_pending():
    if PENDING_FILE.exists():
        try:
            return json.loads(PENDING_FILE.read_text("utf-8"))
        except Exception:
            pass
    return []

def _save_pending(pendientes: list):
    PENDING_FILE.write_text(
        json.dumps(pendientes, indent=2, ensure_ascii=False), "utf-8"
    )

def _add_pending(id_solicitud: str, tipo: str, fecha_ini: str,
                 fecha_fin: str, carp_zip: str, carp_xml: str):
    pendientes = _load_pending()
    # Evitar duplicados
    pendientes = [p for p in pendientes if p.get("id_solicitud") != id_solicitud]
    pendientes.append({
        "id_solicitud": id_solicitud,
        "tipo":         tipo,
        "fecha_ini":    fecha_ini,
        "fecha_fin":    fecha_fin,
        "carp_zip":     carp_zip,
        "carp_xml":     carp_xml,
        "creada":       datetime.datetime.now().isoformat(),
    })
    _save_pending(pendientes)

def _remove_pending(id_solicitud: str):
    pendientes = [p for p in _load_pending()
                  if p.get("id_solicitud") != id_solicitud]
    _save_pending(pendientes)

# ── Config ────────────────────────────────────────────────────
CFG_DEFAULT = {
    "rfc": "", "fiel_cer": "", "fiel_key": "",
    "fiel_cer_nombre": "", "fiel_key_nombre": "",
    "notif_email": "", "notif_cc": "",
    "notif_activo": True,
    "smtp_password": "",        # Contrasena de aplicacion de Gmail (no la normal)
    "dia_auto": 1, "hora_auto": "08:00",
    "nombre": "", "regimen": "",
}

def _load_cfg():
    if CFG_FILE.exists():
        try:
            return {**CFG_DEFAULT, **json.loads(CFG_FILE.read_text("utf-8"))}
        except Exception:
            pass
    return dict(CFG_DEFAULT)

def _save_cfg(cfg):
    CFG_FILE.write_text(json.dumps(cfg, indent=2, ensure_ascii=False), "utf-8")

def _load_hist():
    if HIST_FILE.exists():
        try:
            return json.loads(HIST_FILE.read_text("utf-8"))
        except Exception:
            pass
    return {"primera_ejecucion": True, "ultima_fecha": None,
            "uuids": [], "ejecuciones": []}

def _save_hist(h):
    HIST_FILE.write_text(
        json.dumps(h, indent=2, ensure_ascii=False, default=str), "utf-8"
    )

# ── Envío de correo SMTP ──────────────────────────────────────
def _enviar_correo(ruta_excel: str, fecha_ini: str, fecha_fin: str,
                   total: int, nuevos: int, overwrite: int) -> tuple[bool, str]:
    """
    Envía el reporte Excel por correo via SMTP usando Gmail.
    Requiere una contraseña de aplicación de Google (no la contraseña normal).
    Devuelve (ok, mensaje).
    """
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text      import MIMEText
    from email.mime.base      import MIMEBase
    from email                import encoders

    cfg = _load_cfg()
    destinatario = cfg.get("notif_email", "").strip()
    smtp_password = cfg.get("smtp_password", "").strip()

    if not destinatario:
        return False, "No hay correo de destino configurado."
    if not smtp_password:
        return False, "No hay contrasena de aplicacion configurada."
    if not ruta_excel or not Path(ruta_excel).exists():
        return False, "El archivo Excel no existe."

    periodo = f"{fecha_ini} al {fecha_fin}"
    asunto  = f"Reporte CFDIs — {periodo} — ContaSAT"

    cuerpo_html = f"""
    <html><body style="font-family:Arial,sans-serif;color:#333;max-width:600px">
      <div style="background:#1F3864;padding:20px;border-radius:8px 8px 0 0">
        <h2 style="color:#fff;margin:0">ContaSAT — Reporte de CFDIs</h2>
        <p style="color:#aac4f0;margin:6px 0 0">Descarga automatica completada</p>
      </div>
      <div style="background:#f8f9fa;padding:20px;border:1px solid #dee2e6;border-top:none">
        <p>Se ha completado la descarga de CFDIs para el periodo:</p>
        <p style="font-size:18px;font-weight:bold;color:#1F3864">{periodo}</p>
        <table style="width:100%;border-collapse:collapse;margin:16px 0">
          <tr style="background:#1F3864;color:#fff">
            <th style="padding:10px;text-align:left">Concepto</th>
            <th style="padding:10px;text-align:right">Cantidad</th>
          </tr>
          <tr style="background:#fff">
            <td style="padding:10px;border-bottom:1px solid #dee2e6">Total CFDIs procesados</td>
            <td style="padding:10px;text-align:right;font-weight:bold">{total}</td>
          </tr>
          <tr style="background:#f8f9fa">
            <td style="padding:10px;border-bottom:1px solid #dee2e6">CFDIs nuevos</td>
            <td style="padding:10px;text-align:right;color:#00845A;font-weight:bold">{nuevos}</td>
          </tr>
          <tr style="background:#fff">
            <td style="padding:10px">Sobreescritos (duplicados)</td>
            <td style="padding:10px;text-align:right">{overwrite}</td>
          </tr>
        </table>
        <p style="color:#666;font-size:13px">
          Se adjunta el reporte completo en formato Excel con las hojas
          Resumen, Emitidas y Recibidas.
        </p>
      </div>
      <div style="background:#e9ecef;padding:12px 20px;border-radius:0 0 8px 8px;
                  font-size:12px;color:#666;border:1px solid #dee2e6;border-top:none">
        Generado automaticamente por ContaSAT · {datetime.datetime.now().strftime("%d/%m/%Y %H:%M")}
      </div>
    </body></html>
    """

    try:
        msg = MIMEMultipart("alternative")
        msg["From"]    = destinatario
        msg["To"]      = destinatario
        msg["Subject"] = asunto

        cc = cfg.get("notif_cc", "").strip()
        if cc:
            msg["Cc"] = cc

        msg.attach(MIMEText(cuerpo_html, "html", "utf-8"))

        # Adjuntar Excel
        with open(ruta_excel, "rb") as f:
            adjunto = MIMEBase("application",
                               "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            adjunto.set_payload(f.read())
            encoders.encode_base64(adjunto)
            nombre_archivo = Path(ruta_excel).name
            adjunto.add_header("Content-Disposition", "attachment",
                               filename=nombre_archivo)
            msg.attach(adjunto)

        # Enviar via Gmail SMTP
        destinatarios = [destinatario]
        if cc:
            destinatarios += [c.strip() for c in cc.split(",") if c.strip()]

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(destinatario, smtp_password)
            server.sendmail(destinatario, destinatarios, msg.as_bytes())

        log.info(f"Correo enviado a {destinatarios}")
        return True, f"Correo enviado a {destinatario}"

    except smtplib.SMTPAuthenticationError:
        return False, ("Error de autenticacion Gmail. Verifica que hayas generado "
                       "una 'Contrasena de aplicacion' en myaccount.google.com/security "
                       "y la hayas pegado en Configuracion > Correo.")
    except smtplib.SMTPException as e:
        return False, f"Error SMTP: {e}"
    except Exception as e:
        return False, f"Error al enviar correo: {e}"
NS = {
    "cfdi":  "http://www.sat.gob.mx/cfd/4",
    "cfdi3": "http://www.sat.gob.mx/cfd/3",
    "tfd":   "http://www.sat.gob.mx/TimbreFiscalDigital",
}

# ── Catálogo UsoCFDI del SAT (Anexo 20 CFDI 4.0) ─────────────
# vista: "ingreso" = va a la pestaña Ingresos
#         "gasto"  = va a la pestaña Gastos
#         "neutro" = sin efectos fiscales / pagos
USO_CFDI_CAT = {
    "G01": {"desc": "Adquisición de mercancias",               "vista": "gasto"},
    "G02": {"desc": "Devoluciones, descuentos o bonificaciones","vista": "gasto"},
    "G03": {"desc": "Gastos en general",                       "vista": "gasto"},
    "I01": {"desc": "Construcciones",                          "vista": "gasto"},
    "I02": {"desc": "Mobiliario y equipo de oficina",           "vista": "gasto"},
    "I03": {"desc": "Equipo de transporte",                    "vista": "gasto"},
    "I04": {"desc": "Equipo de computo y accesorios",          "vista": "gasto"},
    "I05": {"desc": "Dados, troqueles y herramental",          "vista": "gasto"},
    "I06": {"desc": "Comunicaciones telefónicas",              "vista": "gasto"},
    "I07": {"desc": "Comunicaciones satelitales",              "vista": "gasto"},
    "I08": {"desc": "Otra maquinaria y equipo",                "vista": "gasto"},
    "D01": {"desc": "Honorarios médicos y gastos hospitalarios","vista": "gasto"},
    "D02": {"desc": "Gastos médicos por incapacidad",          "vista": "gasto"},
    "D03": {"desc": "Gastos funerales",                        "vista": "gasto"},
    "D04": {"desc": "Donativos",                               "vista": "gasto"},
    "D05": {"desc": "Intereses hipotecarios",                  "vista": "gasto"},
    "D06": {"desc": "Aportaciones voluntarias al SAR",         "vista": "gasto"},
    "D07": {"desc": "Primas por seguros de gastos médicos",    "vista": "gasto"},
    "D08": {"desc": "Gastos de transportación escolar",        "vista": "gasto"},
    "D09": {"desc": "Depósitos en cuentas para el ahorro",     "vista": "gasto"},
    "D10": {"desc": "Pagos por servicios educativos",          "vista": "gasto"},
    "CN01":{"desc": "Nómina",                                  "vista": "ingreso"},  # INGRESO: somos el empleador
    "CP01":{"desc": "Pagos",                                   "vista": "neutro"},
    "S01": {"desc": "Sin efectos fiscales",                    "vista": "neutro"},
}

# Claves de impuestos SAT
IMPUESTO_NOMBRE = {
    "001": "ISR",
    "002": "IVA",
    "003": "IEPS",
}


def _parsear(ruta: Path, tipo: str) -> dict:
    """
    Parsea un XML de CFDI 3.3 o 4.0.
    Soporta namespace con y sin prefijo (el SAT usa ambos).
    USO_CFDI_CAT y USO_A_CATEGORIA deben estar definidos antes de llamar.
    """
    try:
        tree = ET.parse(ruta)
        root = tree.getroot()

        # El namespace puede venir con o sin el URI completo
        # Ejemplo: {http://www.sat.gob.mx/cfd/4}Comprobante  o  cfdi:Comprobante
        tag = root.tag
        ns_cfdi = ""
        if "{" in tag:
            ns_cfdi = tag.split("}")[0] + "}"   # e.g. "{http://www.sat.gob.mx/cfd/4}"

        def nodo(nombre):
            """Busca un nodo hijo directo con o sin namespace."""
            # Con namespace completo
            el = root.find(f"{ns_cfdi}{nombre}")
            if el is not None:
                return el
            # Sin namespace (algunos CFDIs lo omiten)
            el = root.find(nombre)
            if el is not None:
                return el
            # Con prefijos registrados en NS
            for pref in ("cfdi", "cfdi3"):
                el = root.find(f"{pref}:{nombre}", NS)
                if el is not None:
                    return el
            return None

        def nodo_deep(path):
            """Busca un nodo en profundidad probando variantes de namespace."""
            for ns_uri in (ns_cfdi, ""):
                partes = path.split("/")
                el = root
                for p in partes:
                    el = el.find(f"{ns_uri}{p}") if el is not None else None
                if el is not None:
                    return el
            # Fallback xpath con prefijos
            for pref in ("cfdi", "cfdi3"):
                path_p = "/".join(f"{pref}:{p}" for p in path.split("/"))
                el = root.find(path_p, NS)
                if el is not None:
                    return el
            return None

        def tfd_find():
            """Busca el TimbreFiscalDigital en cualquier namespace."""
            NS_TFD = "http://www.sat.gob.mx/TimbreFiscalDigital"
            el = root.find(f".//{{{NS_TFD}}}TimbreFiscalDigital")
            if el is not None:
                return el
            return root.find(".//tfd:TimbreFiscalDigital", NS)

        def a(el, attr):
            return el.get(attr, "") if el is not None else ""

        emisor   = nodo("Emisor")
        receptor = nodo("Receptor")
        tfd      = tfd_find()
        concepto = nodo_deep("Conceptos/Concepto")

        # UUID puede estar en el TimbreFiscalDigital o como atributo del Comprobante
        uuid = a(tfd, "UUID") or root.get("Folio", "")

        fecha = root.get("Fecha", "")
        try:
            total = float(root.get("Total", 0) or 0)
        except (ValueError, TypeError):
            total = 0.0

        uso_cfdi      = a(receptor, "UsoCFDI")
        uso_info      = USO_CFDI_CAT.get(uso_cfdi, {})
        uso_cfdi_desc = uso_info.get("desc", "")
        uso_vista     = uso_info.get("vista", "gasto")  # gasto/ingreso/neutro

        # ── Impuestos a nivel Comprobante ────────────────────
        # Estructura: cfdi:Impuestos/cfdi:Retenciones/cfdi:Retencion
        #                           /cfdi:Traslados/cfdi:Traslado
        def _imp_float(val):
            try: return float(val or 0)
            except: return 0.0

        retenciones = []
        traslados   = []

        imp_root = (root.find(f"{ns_cfdi}Impuestos") or
                    root.find("cfdi:Impuestos", NS) or
                    root.find("Impuestos"))
        if imp_root is not None:
            for ret in (imp_root.findall(f"{ns_cfdi}Retenciones/{ns_cfdi}Retencion") or
                        imp_root.findall("cfdi:Retenciones/cfdi:Retencion", NS) or
                        imp_root.findall("Retenciones/Retencion")):
                imp_key = ret.get("Impuesto", "")
                retenciones.append({
                    "impuesto": imp_key,
                    "nombre":   IMPUESTO_NOMBRE.get(imp_key, imp_key),
                    "importe":  _imp_float(ret.get("Importe")),
                })
            for tra in (imp_root.findall(f"{ns_cfdi}Traslados/{ns_cfdi}Traslado") or
                        imp_root.findall("cfdi:Traslados/cfdi:Traslado", NS) or
                        imp_root.findall("Traslados/Traslado")):
                imp_key = tra.get("Impuesto", "")
                traslados.append({
                    "impuesto": imp_key,
                    "nombre":   IMPUESTO_NOMBRE.get(imp_key, imp_key),
                    "tasa":     tra.get("TasaOCuota", ""),
                    "importe":  _imp_float(tra.get("Importe")),
                    "base":     _imp_float(tra.get("Base")),
                })

        # Totales de impuestos (convenientes para el dashboard)
        total_ret_isr = sum(r["importe"] for r in retenciones if r["impuesto"] == "001")
        total_ret_iva = sum(r["importe"] for r in retenciones if r["impuesto"] == "002")
        total_iva_tra = sum(t["importe"] for t in traslados   if t["impuesto"] == "002")
        total_ret_ieps= sum(r["importe"] for r in retenciones if r["impuesto"] == "003")

        # ── Detección de nómina por complemento ──────────────
        NS_NOM12 = "http://www.sat.gob.mx/nomina12"
        es_nomina = (
            uso_cfdi == "CN01" or
            root.find(f".//{{{NS_NOM12}}}Nomina") is not None or
            root.find(".//nomina12:Nomina", {"nomina12": NS_NOM12}) is not None or
            root.get("TipoDeComprobante", "") == "N"
        )
        if es_nomina:
            uso_vista = "ingreso"  # Nómina = somos el empleador = ingreso de la empresa

        # ── Tipo de comprobante final ─────────────────────────
        tipo_comp = root.get("TipoDeComprobante", "")
        # E = Egreso (nota de crédito), las recibidas son favor nuestro
        if tipo_comp == "E" and tipo == "Recibida":
            uso_vista = "ingreso"

        return {
            "tipo":             tipo,
            "uuid":             uuid,
            "fecha":            fecha[:10] if fecha else "",
            "serie":            root.get("Serie", ""),
            "folio":            root.get("Folio", ""),
            "rfc_emisor":       a(emisor,   "Rfc"),
            "nombre_emisor":    a(emisor,   "Nombre"),
            "rfc_receptor":     a(receptor, "Rfc"),
            "nombre_receptor":  a(receptor, "Nombre"),
            "uso_cfdi":         uso_cfdi,
            "uso_cfdi_desc":    uso_cfdi_desc,
            "uso_vista":        uso_vista,
            "es_nomina":        es_nomina,
            "descripcion":      a(concepto, "Descripcion"),
            "subtotal":         root.get("SubTotal", ""),
            "total":            total,
            "moneda":           root.get("Moneda", "MXN"),
            "tipo_comprobante": tipo_comp,
            "metodo_pago":      root.get("MetodoPago", ""),
            "forma_pago":       root.get("FormaPago", ""),
            "regimen_emisor":   a(emisor,   "RegimenFiscal"),
            "clave_prod_serv":  a(concepto, "ClaveProdServ"),
            "archivo":          str(ruta),
            "retenciones":      retenciones,
            "traslados":        traslados,
            "ret_isr":          round(total_ret_isr,  2),
            "ret_iva":          round(total_ret_iva,  2),
            "iva_trasladado":   round(total_iva_tra,  2),
            "ret_ieps":         round(total_ret_ieps, 2),
        }
    except Exception as e:
        log.error(f"Error parseando {ruta}: {e}")
        return {"tipo": tipo, "archivo": str(ruta.name), "error": str(e), "total": 0}

def _emit(msg: str, level: str = "info"):
    log.info(msg)
    _log_lines.append({"msg": msg, "level": level,
                        "ts": datetime.datetime.now().strftime("%H:%M:%S")})
    if len(_log_lines) > 500:
        _log_lines.pop(0)

# ══════════════════════════════════════════════════════════════
#  FLASK
# ══════════════════════════════════════════════════════════════
app = Flask(__name__)

@app.route("/")
def index():
    return send_file(str(GUI_FILE))

@app.route("/<path:filename>")
def static_files(filename):
    return send_from_directory(str(BASE_DIR), filename)

# ── Config ────────────────────────────────────────────────────
@app.route("/api/config", methods=["GET"])
def get_config():
    return jsonify(_load_cfg())

@app.route("/api/config", methods=["POST"])
def save_config():
    try:
        cfg = _load_cfg()
        cfg.update(request.get_json() or {})
        _save_cfg(cfg)
        return jsonify({"ok": True, "msg": "Configuracion guardada."})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})

# ── e.firma ───────────────────────────────────────────────────
@app.route("/api/fiel/status", methods=["GET"])
def fiel_status():
    cfg = _load_cfg()
    return jsonify({
        "cargada":     _signer is not None,
        "cer_nombre":  cfg.get("fiel_cer_nombre", ""),
        "key_nombre":  cfg.get("fiel_key_nombre", ""),
        "cer_ruta":    cfg.get("fiel_cer", ""),
        "key_ruta":    cfg.get("fiel_key", ""),
        "tiene_rutas": bool(cfg.get("fiel_cer") and cfg.get("fiel_key")),
    })

@app.route("/api/fiel/cargar", methods=["POST"])
def cargar_fiel():
    global _signer, _sat
    try:
        d = request.get_json() or {}
        cer_b64  = d.get("cer_b64", "")
        key_b64  = d.get("key_b64", "")
        password = d.get("password", "")
        cer_nombre = d.get("cer_nombre", "")
        key_nombre = d.get("key_nombre", "")
        cer_ruta = d.get("cer_ruta", "")
        key_ruta = d.get("key_ruta", "")

        if cer_ruta and key_ruta and not cer_b64:
            cp = Path(cer_ruta); kp = Path(key_ruta)
            if not cp.exists():
                return jsonify({"ok": False, "msg": f"No se encontro: {cp.name}"})
            if not kp.exists():
                return jsonify({"ok": False, "msg": f"No se encontro: {kp.name}"})
            cer_bytes  = cp.read_bytes()
            key_bytes  = kp.read_bytes()
            cer_nombre = cer_nombre or cp.name
            key_nombre = key_nombre or kp.name
        else:
            if not cer_b64 or not key_b64:
                return jsonify({"ok": False, "msg": "Carga los archivos .cer y .key."})
            cer_bytes = base64.b64decode(cer_b64)
            key_bytes = base64.b64decode(key_b64)

        pwd = password.encode("utf-8") if isinstance(password, str) else password
        _signer = Signer.load(certificate=cer_bytes, key=key_bytes, password=pwd)
        _sat    = SAT(signer=_signer)
        rfc     = _signer.rfc
        log.info(f"Signer cargado. RFC: {rfc}")

        cfg = _load_cfg()
        if cer_ruta: cfg["fiel_cer"] = cer_ruta
        if key_ruta: cfg["fiel_key"] = key_ruta
        cfg.update({"fiel_cer_nombre": cer_nombre,
                    "fiel_key_nombre": key_nombre, "rfc": rfc})
        _save_cfg(cfg)

        return jsonify({"ok": True, "msg": f"e.firma valida. RFC: {rfc}",
                        "rfc": rfc, "cer_nombre": cer_nombre, "key_nombre": key_nombre})
    except Exception as e:
        _signer = None; _sat = None
        msg = str(e)
        if "password" in msg.lower() or "decrypt" in msg.lower():
            msg = "Contrasena incorrecta."
        elif "certificate" in msg.lower():
            msg = "Archivo .cer invalido."
        elif "key" in msg.lower():
            msg = "Archivo .key invalido."
        return jsonify({"ok": False, "msg": f"Error e.firma: {msg}"})

@app.route("/api/fiel/cargar-guardada", methods=["POST"])
def cargar_fiel_guardada():
    cfg      = _load_cfg()
    password = (request.get_json() or {}).get("password", "")
    with app.test_request_context(
        "/api/fiel/cargar", method="POST",
        json={"cer_ruta": cfg.get("fiel_cer",""),
              "key_ruta": cfg.get("fiel_key",""),
              "password": password},
        content_type="application/json"
    ):
        return cargar_fiel()

# ── Rango ─────────────────────────────────────────────────────
@app.route("/api/rango/automatico", methods=["GET"])
def rango_automatico():
    h   = _load_hist()
    hoy = datetime.date.today()
    if h["primera_ejecucion"] or not h["ultima_fecha"]:
        ini  = datetime.date(hoy.year, 1, 1)
        modo = "primera"
    else:
        ultima = datetime.date.fromisoformat(h["ultima_fecha"])
        ini    = ultima - datetime.timedelta(days=1)
        modo   = "incremental"
    return jsonify({"inicio": str(ini), "fin": str(hoy),
                    "modo": modo, "primera_ejecucion": h["primera_ejecucion"]})

# ── Historial ─────────────────────────────────────────────────
@app.route("/api/historial", methods=["GET"])
def get_historial():
    return jsonify(_load_hist())

@app.route("/api/historial/limpiar", methods=["POST"])
def limpiar_historial():
    try:
        _save_hist({"primera_ejecucion": True, "ultima_fecha": None,
                    "uuids": [], "ejecuciones": []})
        return jsonify({"ok": True, "msg": "Historial limpiado."})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})

# ── Correo SMTP ───────────────────────────────────────────────
@app.route("/api/correo/probar", methods=["POST"])
def probar_correo():
    """Envía un correo de prueba para validar la configuración SMTP."""
    try:
        import smtplib
        cfg = _load_cfg()
        destinatario  = cfg.get("notif_email", "").strip()
        smtp_password = cfg.get("smtp_password", "").strip()

        if not destinatario:
            return jsonify({"ok": False, "msg": "Configura el correo de destino primero."})
        if not smtp_password:
            return jsonify({"ok": False, "msg": "Configura la contrasena de aplicacion primero."})

        from email.mime.text import MIMEText
        msg = MIMEText(
            "Este es un correo de prueba enviado por ContaSAT.\n\n"
            "Si lo recibes, la configuracion SMTP es correcta.",
            "plain", "utf-8"
        )
        msg["From"]    = destinatario
        msg["To"]      = destinatario
        msg["Subject"] = "ContaSAT — Prueba de correo"

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(destinatario, smtp_password)
            server.sendmail(destinatario, [destinatario], msg.as_bytes())

        return jsonify({"ok": True, "msg": f"Correo de prueba enviado a {destinatario}. Revisa tu bandeja."})

    except smtplib.SMTPAuthenticationError:
        return jsonify({"ok": False,
                        "msg": "Error de autenticacion. Verifica que sea una "
                               "Contrasena de aplicacion de Google, no tu contrasena normal."})
    except Exception as e:
        return jsonify({"ok": False, "msg": f"Error: {e}"})

# ── Solicitudes pendientes ────────────────────────────────────
@app.route("/api/pendientes", methods=["GET"])
def get_pendientes():
    return jsonify({"pendientes": _load_pending()})

@app.route("/api/pendientes/reanudar", methods=["POST"])
def reanudar_pendiente():
    global _descargando
    if not _sat:
        return jsonify({"ok": False, "msg": "Carga tu e.firma primero."})
    if _descargando:
        return jsonify({"ok": False, "msg": "Ya hay una descarga en proceso."})
    d            = request.get_json() or {}
    id_solicitud = d.get("id_solicitud")
    pendientes   = _load_pending()
    solicitud    = next((p for p in pendientes if p["id_solicitud"] == id_solicitud), None)
    if not solicitud:
        return jsonify({"ok": False, "msg": "Solicitud no encontrada."})
    threading.Thread(
        target=_reanudar_worker,
        args=(solicitud,),
        daemon=True,
    ).start()
    return jsonify({"ok": True, "msg": f"Reanudando solicitud {id_solicitud}..."})

@app.route("/api/pendientes/cancelar", methods=["POST"])
def cancelar_pendiente():
    id_solicitud = (request.get_json() or {}).get("id_solicitud")
    _remove_pending(id_solicitud)
    return jsonify({"ok": True, "msg": "Solicitud eliminada del registro."})

@app.route("/api/plan/estado", methods=["GET"])
def get_plan_estado():
    """Devuelve el plan de descarga guardado si existe."""
    plan_file = DATA_DIR / "plan_descarga.json"
    if not plan_file.exists():
        return jsonify({"tiene_plan": False})
    try:
        plan = json.loads(plan_file.read_text("utf-8"))
        pendientes = [p for p in plan if p["estado"] in ("pendiente", "en_proceso")]
        completados = [p for p in plan if p["estado"] == "completado"]
        return jsonify({
            "tiene_plan":   True,
            "total":        len(plan),
            "pendientes":   len(pendientes),
            "completados":  len(completados),
            "bloques":      plan,
        })
    except Exception as e:
        return jsonify({"tiene_plan": False, "error": str(e)})

@app.route("/api/plan/reanudar", methods=["POST"])
def reanudar_plan():
    """Reanuda el plan completo desde el primer bloque no completado."""
    global _descargando
    if not _sat:
        return jsonify({"ok": False, "msg": "Carga tu e.firma primero."})
    if _descargando:
        return jsonify({"ok": False, "msg": "Ya hay una descarga en proceso."})
    plan_file = DATA_DIR / "plan_descarga.json"
    if not plan_file.exists():
        return jsonify({"ok": False, "msg": "No hay plan guardado."})
    plan = json.loads(plan_file.read_text("utf-8"))
    threading.Thread(
        target=_reanudar_plan_worker,
        args=(plan, plan_file),
        daemon=True,
    ).start()
    return jsonify({"ok": True, "msg": "Reanudando plan de descarga..."})

@app.route("/api/plan/cancelar", methods=["POST"])
def cancelar_plan():
    plan_file = DATA_DIR / "plan_descarga.json"
    if plan_file.exists():
        plan_file.unlink()
    return jsonify({"ok": True, "msg": "Plan cancelado."})

def _reanudar_plan_worker(plan: list, plan_file: Path):
    """Reanuda todos los bloques pendientes o en_proceso del plan."""
    global _progreso, _descargando
    _descargando = True
    _progreso    = 0
    _log_lines.clear()

    hist   = _load_hist()
    uuids  = set(hist.get("uuids", []))
    todos  = []
    nuevos = overwr = 0

    bloques_pendientes = [p for p in plan if p["estado"] in ("pendiente", "en_proceso")]
    total = len(bloques_pendientes)
    _emit(f"Reanudando plan: {total} bloque(s) pendiente(s) de {len(plan)} totales.", "info")

    try:
        from satcfdi.pacs.sat import EstadoComprobante
        estado_comp = EstadoComprobante.VIGENTE
    except ImportError:
        estado_comp = "1"

    import time

    def _guardar_plan():
        plan_file.write_text(json.dumps(plan, indent=2, ensure_ascii=False), "utf-8")

    try:
        for idx, item in enumerate(bloques_pendientes, 1):
            tipo       = item["tipo"]
            bloque_ini = datetime.date.fromisoformat(item["fecha_ini"])
            bloque_fin = datetime.date.fromisoformat(item["fecha_fin"])
            carp_zip   = Path(item["carp_zip"])
            carp_xml   = Path(item["carp_xml"])
            carp_zip.mkdir(parents=True, exist_ok=True)
            carp_xml.mkdir(parents=True, exist_ok=True)

            _emit(f"[{idx}/{total}] {tipo.upper()} {bloque_ini} -> {bloque_fin}...", "info")
            item["estado"] = "en_proceso"
            _guardar_plan()

            # Si ya tiene ID de solicitud, retomar verificacion
            id_solicitud = item.get("id_solicitud")
            paquetes = []

            if id_solicitud:
                _emit(f"Retomando solicitud existente: {id_solicitud}", "info")
            else:
                # Crear nueva solicitud para este bloque
                try:
                    if tipo == "recibidas":
                        resp = _sat.recover_comprobante_received_request(
                            fecha_inicial      = bloque_ini,
                            fecha_final        = bloque_fin,
                            rfc_receptor       = _signer.rfc,
                            tipo_solicitud     = TipoDescargaMasivaTerceros.CFDI,
                            estado_comprobante = estado_comp,
                        )
                    else:
                        resp = _sat.recover_comprobante_emitted_request(
                            fecha_inicial      = bloque_ini,
                            fecha_final        = bloque_fin,
                            rfc_emisor         = _signer.rfc,
                            tipo_solicitud     = TipoDescargaMasivaTerceros.CFDI,
                            estado_comprobante = estado_comp,
                        )
                    id_solicitud = resp.get("IdSolicitud")
                    if not id_solicitud:
                        _emit(f"Error SAT: {resp.get('Mensaje','')}", "error")
                        item["estado"] = "error"
                        _guardar_plan()
                        continue
                    item["id_solicitud"] = id_solicitud
                    _guardar_plan()
                    _emit(f"Nueva solicitud: {id_solicitud}", "ok")
                except Exception as re:
                    _emit(f"Error al solicitar: {re}", "error")
                    item["estado"] = "error"
                    _guardar_plan()
                    continue

            # Verificar hasta que este lista
            for intento in range(200):
                time.sleep(30 if intento > 0 else 5)
                status   = _sat.recover_comprobante_status(id_solicitud)
                estado   = status.get("EstadoSolicitud", "")
                paquetes = status.get("IdsPaquetes", [])
                n_cfdis  = status.get("NumeroCFDIs", 0)
                _emit(f"Verificando... estado={estado} paquetes={len(paquetes)} cfdis={n_cfdis}", "info")
                _progreso = int((idx - 1 + 0.5) / total * 90)

                if str(estado) == "3":
                    _emit(f"Listo. {len(paquetes)} paquete(s).", "ok")
                    break
                elif str(estado) == "5":
                    _emit("Sin CFDIs en este bloque.", "warn")
                    paquetes = []
                    break
                elif str(estado) in ["1", "2"]:
                    continue
                else:
                    _emit(f"Estado inesperado: {estado}", "error")
                    paquetes = []
                    break

            # Descargar paquetes
            for i, id_paquete in enumerate(paquetes):
                _emit(f"Descargando paquete {i+1}/{len(paquetes)}: {id_paquete}", "info")
                try:
                    resp_dict, contenido_b64 = _sat.recover_comprobante_download(id_paquete)
                    if not contenido_b64:
                        continue
                    data = base64.b64decode(contenido_b64)
                    ruta_zip = carp_zip / f"{id_paquete}.zip"
                    ruta_zip.write_bytes(data)
                    with zipfile.ZipFile(ruta_zip) as z:
                        for nombre in [x for x in z.namelist() if x.lower().endswith(".xml")]:
                            contenido = z.read(nombre)
                            ruta_xml  = carp_xml / nombre
                            ruta_xml.write_bytes(contenido)
                            label = "Emitida" if tipo == "emitidas" else "Recibida"
                            cfdi  = _parsear(ruta_xml, label)
                            uuid  = cfdi.get("uuid", "")
                            if uuid:
                                if uuid in uuids: overwr += 1
                                else: uuids.add(uuid); nuevos += 1
                            todos.append(cfdi)
                except Exception as de:
                    _emit(f"Error paquete {id_paquete}: {de}", "error")

            item["estado"] = "completado"
            _guardar_plan()
            _remove_pending(id_solicitud)
            _progreso = int(idx / total * 90)
            _emit(f"Bloque {idx}/{total} completado.", "ok")

        # Reporte final
        if todos:
            _emit("Generando reporte Excel...", "info")
            hoy = datetime.date.today()
            _generar_excel(todos, DATA_DIR / "reportes", hoy.replace(day=1), hoy)

        hist["primera_ejecucion"] = False
        hist["ultima_fecha"]      = str(datetime.date.today())
        hist["uuids"]             = list(uuids)
        hist["ejecuciones"].append({
            "fecha":  datetime.datetime.now().isoformat(),
            "inicio": plan[0]["fecha_ini"] if plan else "",
            "fin":    plan[-1]["fecha_fin"] if plan else "",
            "total":  len(todos), "nuevos": nuevos, "overwrite": overwr,
        })
        _save_hist(hist)

        # Si todos los bloques estan completados, eliminar el plan
        if all(p["estado"] == "completado" for p in plan):
            plan_file.unlink(missing_ok=True)

        _progreso = 100
        _emit(f"COMPLETADO: {len(todos)} CFDIs | {nuevos} nuevos | {overwr} overwrite", "ok")

    except Exception as e:
        _emit(f"Error inesperado: {e}", "error")
        log.exception("Error en reanudar plan worker")
    finally:
        _descargando = False

def _dividir_en_trimestres(fecha_ini: datetime.date,
                            fecha_fin: datetime.date) -> list:
    """
    El SAT no acepta rangos mayores a 3 meses por solicitud.
    Divide cualquier rango en bloques de maximo 3 meses.
    """
    bloques = []
    actual  = fecha_ini
    while actual <= fecha_fin:
        # Fin del bloque = 3 meses despues - 1 dia, sin exceder fecha_fin
        mes_fin = actual.month + 2          # +2 porque vamos a fin del 3er mes
        anio_fin = actual.year + (mes_fin - 1) // 12
        mes_fin  = ((mes_fin - 1) % 12) + 1
        # Ultimo dia del mes de fin
        import calendar
        ultimo_dia = calendar.monthrange(anio_fin, mes_fin)[1]
        fin_bloque = datetime.date(anio_fin, mes_fin, ultimo_dia)
        fin_bloque = min(fin_bloque, fecha_fin)
        bloques.append((actual, fin_bloque))
        actual = fin_bloque + datetime.timedelta(days=1)
    return bloques
    """
    Retoma una solicitud ya enviada al SAT sin volver a crearla.
    Solo verifica el estado y descarga los paquetes cuando esten listos.
    """
    global _progreso, _descargando
    _descargando = True
    _progreso    = 0
    _log_lines.clear()

    id_solicitud = solicitud["id_solicitud"]
    tipo         = solicitud["tipo"]
    fecha_ini    = solicitud["fecha_ini"]
    fecha_fin    = solicitud["fecha_fin"]
    carp_zip     = Path(solicitud["carp_zip"])
    carp_xml     = Path(solicitud["carp_xml"])
    carp_zip.mkdir(parents=True, exist_ok=True)
    carp_xml.mkdir(parents=True, exist_ok=True)

    hist   = _load_hist()
    uuids  = set(hist.get("uuids", []))
    todos  = []
    nuevos = overwr = 0

    _emit(f"Reanudando solicitud {tipo.upper()} — ID: {id_solicitud}", "info")
    _emit(f"Período: {fecha_ini} -> {fecha_fin}", "info")

    try:
        import time
        paquetes = []
        for intento in range(200):
            if intento > 0:
                time.sleep(30)
            else:
                time.sleep(3)

            status   = _sat.recover_comprobante_status(id_solicitud)
            estado   = status.get("EstadoSolicitud", "")
            paquetes = status.get("IdsPaquetes", [])
            n_cfdis  = status.get("NumeroCFDIs", 0)
            _emit(f"Intento {intento+1}: estado={estado} paquetes={len(paquetes)} cfdis={n_cfdis}", "info")
            _progreso = min(10 + intento * 2, 60)

            if str(estado) == "3":
                _emit(f"Solicitud lista. {len(paquetes)} paquete(s).", "ok")
                break
            elif str(estado) == "5":
                _emit("Sin CFDIs en el periodo.", "warn")
                _remove_pending(id_solicitud)
                _descargando = False
                return
            elif str(estado) in ["1", "2"]:
                continue
            else:
                _emit(f"Estado inesperado: {estado}", "error")
                _descargando = False
                return

        # Descargar paquetes
        n = 0
        for i, id_paquete in enumerate(paquetes):
            _emit(f"Descargando paquete {i+1}/{len(paquetes)}: {id_paquete}", "info")
            try:
                resp_dict, contenido_b64 = _sat.recover_comprobante_download(id_paquete)
                if not contenido_b64:
                    _emit(f"Paquete vacio: {id_paquete}", "warn")
                    continue
                data = base64.b64decode(contenido_b64)
                ruta_zip = carp_zip / f"{id_paquete}.zip"
                ruta_zip.write_bytes(data)
                n += 1
                with zipfile.ZipFile(ruta_zip) as z:
                    for nombre in [x for x in z.namelist() if x.lower().endswith(".xml")]:
                        contenido = z.read(nombre)
                        ruta_xml  = carp_xml / nombre
                        ruta_xml.write_bytes(contenido)
                        label = "Emitida" if tipo == "emitidas" else "Recibida"
                        cfdi  = _parsear(ruta_xml, label)
                        uuid  = cfdi.get("uuid", "")
                        if uuid:
                            if uuid in uuids: overwr += 1
                            else: uuids.add(uuid); nuevos += 1
                        todos.append(cfdi)
                _progreso = int(60 + (i+1) / max(len(paquetes),1) * 35)
            except Exception as de:
                _emit(f"Error paquete {id_paquete}: {de}", "error")

        if todos:
            _emit("Generando reporte Excel...", "info")
            base = carp_zip.parent.parent
            _generar_excel(todos, base / "reportes",
                           datetime.date.fromisoformat(fecha_ini),
                           datetime.date.fromisoformat(fecha_fin))

        hist["primera_ejecucion"] = False
        hist["ultima_fecha"]      = fecha_fin
        hist["uuids"]             = list(uuids)
        hist["ejecuciones"].append({
            "fecha": datetime.datetime.now().isoformat(),
            "inicio": fecha_ini, "fin": fecha_fin,
            "total": len(todos), "nuevos": nuevos, "overwrite": overwr,
        })
        _save_hist(hist)
        _remove_pending(id_solicitud)
        _progreso = 100
        _emit(f"COMPLETADO: {len(todos)} CFDIs | {nuevos} nuevos | {overwr} overwrite", "ok")

    except Exception as e:
        _emit(f"Error inesperado: {e}", "error")
        log.exception("Error en reanudar worker")
    finally:
        _descargando = False

# ── Descarga ──────────────────────────────────────────────────
@app.route("/api/descarga/iniciar", methods=["POST"])
def iniciar_descarga():
    global _descargando
    if not _sat:
        return jsonify({"ok": False, "msg": "Carga y valida tu e.firma primero."})
    if _descargando:
        return jsonify({"ok": False, "msg": "Ya hay una descarga en proceso."})
    d = request.get_json() or {}
    threading.Thread(
        target=_descarga_worker,
        args=(d.get("inicio"), d.get("fin"), d.get("tipo_cfdi", "ambas")),
        daemon=True,
    ).start()
    return jsonify({"ok": True, "msg": "Descarga iniciada en segundo plano."})

def _descarga_worker(inicio: str, fin: str, tipo_cfdi: str):
    global _progreso, _descargando
    _descargando = True
    _progreso    = 0
    _log_lines.clear()

    fecha_ini = datetime.date.fromisoformat(inicio)
    fecha_fin = datetime.date.fromisoformat(fin)
    hist      = _load_hist()
    uuids     = set(hist.get("uuids", []))
    todos     = []
    nuevos = overwr = 0

    tipos = []
    if tipo_cfdi in ("ambas", "recibidas"): tipos.append("recibidas")
    if tipo_cfdi in ("ambas", "emitidas"):  tipos.append("emitidas")

    bloques     = _dividir_en_trimestres(fecha_ini, fecha_fin)
    total_pasos = len(tipos) * len(bloques)
    paso_actual = 0

    _emit(f"Periodo: {fecha_ini} -> {fecha_fin}", "info")
    _emit(f"Dividido en {len(bloques)} bloque(s) x {len(tipos)} tipo(s) = {total_pasos} solicitudes.", "info")

    # Guardar el plan completo en disco desde el inicio
    # Cada bloque pendiente se marca con estado "pendiente"
    plan_file = DATA_DIR / "plan_descarga.json"
    plan = []
    for tipo in tipos:
        for bi, bf in bloques:
            base     = DATA_DIR / str(bi.year) / f"{bi.month:02d}"
            carp_zip = str(base / tipo / "zips")
            carp_xml = str(base / tipo / "xml")
            plan.append({
                "tipo":      tipo,
                "fecha_ini": str(bi),
                "fecha_fin": str(bf),
                "carp_zip":  carp_zip,
                "carp_xml":  carp_xml,
                "estado":    "pendiente",   # pendiente | en_proceso | completado | error
                "id_solicitud": None,
            })
    plan_file.write_text(json.dumps(plan, indent=2, ensure_ascii=False), "utf-8")

    try:
        from satcfdi.pacs.sat import EstadoComprobante
        estado_comp = EstadoComprobante.VIGENTE
    except ImportError:
        estado_comp = "1"

    import time

    def _guardar_plan():
        plan_file.write_text(json.dumps(plan, indent=2, ensure_ascii=False), "utf-8")

    try:
        for item in plan:
            paso_actual += 1
            tipo      = item["tipo"]
            bloque_ini = datetime.date.fromisoformat(item["fecha_ini"])
            bloque_fin = datetime.date.fromisoformat(item["fecha_fin"])
            carp_zip   = Path(item["carp_zip"])
            carp_xml   = Path(item["carp_xml"])
            carp_zip.mkdir(parents=True, exist_ok=True)
            carp_xml.mkdir(parents=True, exist_ok=True)

            item["estado"] = "en_proceso"
            _guardar_plan()

            _emit(f"[{paso_actual}/{total_pasos}] {tipo.upper()} {bloque_ini} -> {bloque_fin}...", "info")

            # ── Solicitar ────────────────────────────────────────
            try:
                if tipo == "recibidas":
                    resp = _sat.recover_comprobante_received_request(
                        fecha_inicial      = bloque_ini,
                        fecha_final        = bloque_fin,
                        rfc_receptor       = _signer.rfc,
                        tipo_solicitud     = TipoDescargaMasivaTerceros.CFDI,
                        estado_comprobante = estado_comp,
                    )
                else:
                    resp = _sat.recover_comprobante_emitted_request(
                        fecha_inicial      = bloque_ini,
                        fecha_final        = bloque_fin,
                        rfc_emisor         = _signer.rfc,
                        tipo_solicitud     = TipoDescargaMasivaTerceros.CFDI,
                        estado_comprobante = estado_comp,
                    )
            except Exception as re:
                _emit(f"Error al solicitar: {re}", "error")
                item["estado"] = "error"
                _guardar_plan()
                continue

            id_solicitud = resp.get("IdSolicitud")
            cod          = resp.get("CodEstatus", "")
            if not id_solicitud:
                _emit(f"Error SAT {cod}: {resp.get('Mensaje','')}", "error")
                item["estado"] = "error"
                _guardar_plan()
                continue

            item["id_solicitud"] = id_solicitud
            _guardar_plan()
            _add_pending(id_solicitud, tipo, str(bloque_ini),
                         str(bloque_fin), str(carp_zip), str(carp_xml))
            _emit(f"ID: {id_solicitud} | Codigo: {cod}", "ok")

            # ── Verificar ────────────────────────────────────────
            paquetes = []
            for intento in range(200):
                time.sleep(30 if intento > 0 else 5)
                status   = _sat.recover_comprobante_status(id_solicitud)
                estado   = status.get("EstadoSolicitud", "")
                paquetes = status.get("IdsPaquetes", [])
                n_cfdis  = status.get("NumeroCFDIs", 0)
                _emit(f"Verificando... estado={estado} paquetes={len(paquetes)} cfdis={n_cfdis}", "info")
                _progreso = int((paso_actual - 1 + 0.5) / total_pasos * 90)

                if str(estado) == "3":
                    _emit(f"Listo. {len(paquetes)} paquete(s).", "ok")
                    break
                elif str(estado) == "5":
                    _emit("Sin CFDIs en este bloque.", "warn")
                    paquetes = []
                    break
                elif str(estado) in ["1", "2"]:
                    continue
                else:
                    _emit(f"Estado inesperado: {estado}", "error")
                    paquetes = []
                    break

            # ── Descargar ────────────────────────────────────────
            n = 0
            for i, id_paquete in enumerate(paquetes):
                _emit(f"Descargando paquete {i+1}/{len(paquetes)}: {id_paquete}", "info")
                try:
                    resp_dict, contenido_b64 = _sat.recover_comprobante_download(id_paquete)
                    if not contenido_b64:
                        _emit(f"Paquete vacio: {id_paquete}", "warn")
                        continue
                    data = base64.b64decode(contenido_b64)
                    ruta_zip = carp_zip / f"{id_paquete}.zip"
                    ruta_zip.write_bytes(data)
                    n += 1
                    with zipfile.ZipFile(ruta_zip) as z:
                        for nombre in [x for x in z.namelist() if x.lower().endswith(".xml")]:
                            contenido = z.read(nombre)
                            ruta_xml  = carp_xml / nombre
                            ruta_xml.write_bytes(contenido)
                            label = "Emitida" if tipo == "emitidas" else "Recibida"
                            cfdi  = _parsear(ruta_xml, label)
                            uuid  = cfdi.get("uuid", "")
                            if uuid:
                                if uuid in uuids: overwr += 1
                                else: uuids.add(uuid); nuevos += 1
                            todos.append(cfdi)
                except Exception as de:
                    _emit(f"Error paquete {id_paquete}: {de}", "error")

            if n > 0:
                _emit(f"Bloque completado: {n} paquete(s).", "ok")

            item["estado"] = "completado"
            _guardar_plan()
            _remove_pending(id_solicitud)
            _progreso = int(paso_actual / total_pasos * 90)

        # ── Reporte y guardar ────────────────────────────────────
        if todos:
            _emit("Generando reporte Excel...", "info")
            base_rep = DATA_DIR / str(fecha_ini.year) / f"{fecha_ini.month:02d}"
            _generar_excel(todos, base_rep / "reportes", fecha_ini, fecha_fin)

        hist["primera_ejecucion"] = False
        hist["ultima_fecha"]      = str(fecha_fin)
        hist["uuids"]             = list(uuids)
        hist["ejecuciones"].append({
            "fecha":  datetime.datetime.now().isoformat(),
            "inicio": str(fecha_ini), "fin": str(fecha_fin),
            "total":  len(todos), "nuevos": nuevos, "overwrite": overwr,
        })
        _save_hist(hist)
        # Plan completado — eliminar archivo
        if plan_file.exists():
            plan_file.unlink()
        _progreso = 100
        _emit(f"COMPLETADO: {len(todos)} CFDIs | {nuevos} nuevos | {overwr} overwrite", "ok")

        # Enviar correo si esta configurado
        cfg = _load_cfg()
        if cfg.get("notif_activo", True) and cfg.get("notif_email") and cfg.get("smtp_password"):
            _emit("Enviando reporte por correo...", "info")
            ruta_excel_str = str(base_rep / "reportes" / f"CFDIs_{fecha_ini}_{fecha_fin}.xlsx")
            ok, msg = _enviar_correo(ruta_excel_str, str(fecha_ini), str(fecha_fin),
                                     len(todos), nuevos, overwr)
            _emit(msg, "ok" if ok else "warn")

    except Exception as e:
        _emit(f"Error inesperado: {e}", "error")
        log.exception("Error en descarga worker")
    finally:
        _descargando = False

    try:
        for paso, tipo in enumerate(tipos):
            carp_zip = base / tipo / "zips"
            carp_xml = base / tipo / "xml"
            carp_zip.mkdir(parents=True, exist_ok=True)
            carp_xml.mkdir(parents=True, exist_ok=True)

            _emit(f"Solicitando CFDIs {tipo.upper()} {fecha_ini} -> {fecha_fin}...", "info")

            # ── Paso 1: Solicitar ────────────────────────────────
            # Firma real:
            # recover_comprobante_received_request(fecha_inicial, fecha_final,
            #   rfc_receptor=None, rfc_emisor=None,
            #   tipo_solicitud=CFDI, tipo_comprobante=None,
            #   estado_comprobante=None, ...)
            # EstadoComprobante.VIGENTE excluye cancelados (resuelve error 301)
            try:
                from satcfdi.pacs.sat import EstadoComprobante
                estado_comp = EstadoComprobante.VIGENTE
            except ImportError:
                estado_comp = "1"   # "1" = Vigente segun SAT

            if tipo == "recibidas":
                resp = _sat.recover_comprobante_received_request(
                    fecha_inicial       = fecha_ini,
                    fecha_final         = fecha_fin,
                    rfc_receptor        = _signer.rfc,
                    tipo_solicitud      = TipoDescargaMasivaTerceros.CFDI,
                    estado_comprobante  = estado_comp,
                )
            else:
                resp = _sat.recover_comprobante_emitted_request(
                    fecha_inicial       = fecha_ini,
                    fecha_final         = fecha_fin,
                    rfc_emisor          = _signer.rfc,
                    tipo_solicitud      = TipoDescargaMasivaTerceros.CFDI,
                    estado_comprobante  = estado_comp,
                )

            _emit(f"Respuesta SAT: {resp}", "info")
            id_solicitud = resp.get("IdSolicitud")
            cod          = resp.get("CodEstatus", "")

            if not id_solicitud:
                _emit(f"Error SAT codigo {cod}: {resp.get('Mensaje', 'Sin mensaje')}", "error")
                continue

            _emit(f"ID solicitud: {id_solicitud} | Codigo: {cod}", "ok")

            # Guardar solicitud en disco para poder reanudar si se cierra la app
            _add_pending(id_solicitud, tipo, str(fecha_ini), str(fecha_fin),
                         str(carp_zip), str(carp_xml))

            # ── Paso 2: Verificar ────────────────────────────────
            _emit("Verificando estado de la solicitud...", "info")
            import time
            paquetes = []
            for intento in range(200):
                if intento > 0:
                    time.sleep(30)
                else:
                    time.sleep(5)

                status  = _sat.recover_comprobante_status(id_solicitud)
                estado  = status.get("EstadoSolicitud", "")
                paquetes = status.get("IdsPaquetes", [])
                n_cfdis  = status.get("NumeroCFDIs", 0)
                _emit(f"Intento {intento+1}: estado={estado} paquetes={len(paquetes)} cfdis={n_cfdis}", "info")

                if str(estado) == "3":   # Terminada
                    _emit(f"Solicitud lista. {len(paquetes)} paquete(s).", "ok")
                    break
                elif str(estado) == "5":  # Sin comprobantes
                    _emit("Sin CFDIs en el periodo indicado.", "warn")
                    paquetes = []
                    break
                elif str(estado) in ["1", "2"]:  # En proceso
                    _progreso = int((paso + 0.3) / len(tipos) * 80)
                    continue
                else:
                    _emit(f"Estado inesperado: {estado} | {status}", "error")
                    paquetes = []
                    break

            # ── Paso 3: Descargar ────────────────────────────────
            n = 0
            for i, id_paquete in enumerate(paquetes):
                _emit(f"Descargando paquete {i+1}/{len(paquetes)}: {id_paquete}", "info")
                try:
                    # Firma real: recover_comprobante_download(id_paquete) -> (dict, str)
                    resp_dict, contenido_b64 = _sat.recover_comprobante_download(id_paquete)

                    if not contenido_b64:
                        _emit(f"Paquete vacio: {id_paquete}", "warn")
                        continue

                    data     = base64.b64decode(contenido_b64)
                    ruta_zip = carp_zip / f"{id_paquete}.zip"
                    ruta_zip.write_bytes(data)
                    n += 1

                    with zipfile.ZipFile(ruta_zip) as z:
                        for nombre in [x for x in z.namelist() if x.lower().endswith(".xml")]:
                            contenido = z.read(nombre)
                            ruta_xml  = carp_xml / nombre
                            ruta_xml.write_bytes(contenido)
                            label = "Emitida" if tipo == "emitidas" else "Recibida"
                            cfdi  = _parsear(ruta_xml, label)
                            uuid  = cfdi.get("uuid", "")
                            if uuid:
                                if uuid in uuids: overwr += 1
                                else: uuids.add(uuid); nuevos += 1
                            todos.append(cfdi)

                    _progreso = int((paso + 0.6 + (i+1) / max(len(paquetes),1) * 0.4) / len(tipos) * 90)

                except Exception as de:
                    _emit(f"Error descargando paquete {id_paquete}: {de}", "error")
                    log.exception(f"Error paquete {id_paquete}")

            if n > 0:
                _emit(f"{tipo.upper()}: {n} paquete(s), {len([d for d in todos if d.get('tipo')==('Emitida' if tipo=='emitidas' else 'Recibida')])} CFDIs.", "ok")
            # Solicitud procesada — eliminar de pendientes
            _remove_pending(id_solicitud)

        if todos:
            _emit("Generando reporte Excel...", "info")
            _generar_excel(todos, base / "reportes", fecha_ini, fecha_fin)

        hist["primera_ejecucion"]  = False
        hist["ultima_fecha"]       = str(fecha_fin)
        hist["uuids"]              = list(uuids)
        hist["ejecuciones"].append({
            "fecha": datetime.datetime.now().isoformat(),
            "inicio": str(fecha_ini), "fin": str(fecha_fin),
            "total": len(todos), "nuevos": nuevos, "overwrite": overwr,
        })
        _save_hist(hist)
        _progreso = 100
        _emit(f"COMPLETADO: {len(todos)} CFDIs | {nuevos} nuevos | {overwr} overwrite", "ok")

    except Exception as e:
        _emit(f"Error inesperado: {e}", "error")
        log.exception("Error en descarga worker")
    finally:
        _descargando = False

@app.route("/api/descarga/estado", methods=["GET"])
def estado_descarga():
    desde = int(request.args.get("desde", 0))
    return jsonify({
        "progreso":    _progreso,
        "descargando": _descargando,
        "log":         _log_lines[desde:],
        "total_log":   len(_log_lines),
    })

# ── Facturas ─────────────────────────────────────────────────
@app.route("/api/facturas", methods=["GET"])
def get_facturas():
    try:
        tipo = request.args.get("tipo", "todas")
        anio = request.args.get("anio")
        mes  = request.args.get("mes")
        resultados = []
        for xml in DATA_DIR.rglob("*.xml"):
            partes = xml.parts
            if "emitidas"   in partes: t = "Emitida"
            elif "recibidas" in partes: t = "Recibida"
            else: t = "Desconocida"
            if tipo == "emitidas"  and t != "Emitida":  continue
            if tipo == "recibidas" and t != "Recibida": continue
            cfdi = _parsear(xml, t)
            if anio and cfdi.get("fecha","")[:4]  != str(anio):         continue
            if mes  and cfdi.get("fecha","")[5:7] != f"{int(mes):02d}": continue
            resultados.append(cfdi)
        return jsonify({"ok": True, "data": resultados, "total": len(resultados)})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e), "data": []})

@app.route("/api/dashboard/stats", methods=["GET"])
def dashboard_stats():
    try:
        xmls      = list(DATA_DIR.rglob("*.xml"))
        emitidas  = [_parsear(x,"Emitida")  for x in xmls if "emitidas"  in x.parts]
        recibidas = [_parsear(x,"Recibida") for x in xmls if "recibidas" in x.parts]
        emitidas  = [f for f in emitidas  if "error" not in f]
        recibidas = [f for f in recibidas if "error" not in f]
        hist      = _load_hist()
        return jsonify({
            "total_emitidas":  len(emitidas),
            "total_recibidas": len(recibidas),
            "monto_emitido":   round(sum(f.get("total",0) for f in emitidas),  2),
            "monto_recibido":  round(sum(f.get("total",0) for f in recibidas), 2),
            "balance":         round(sum(f.get("total",0) for f in emitidas) -
                                     sum(f.get("total",0) for f in recibidas), 2),
            "ultima_descarga": hist.get("ultima_fecha","Sin descargas"),
            "total_uuids":     len(hist.get("uuids",[])),
        })
    except Exception as e:
        return jsonify({"error": str(e)})

@app.route("/api/reporte/excel", methods=["POST"])
def generar_reporte():
    try:
        xmls  = list(DATA_DIR.rglob("*.xml"))
        datos = ([_parsear(x,"Emitida")  for x in xmls if "emitidas"  in x.parts] +
                 [_parsear(x,"Recibida") for x in xmls if "recibidas" in x.parts])
        datos = [f for f in datos if "error" not in f]
        hoy   = datetime.date.today()
        ruta  = _generar_excel(datos, DATA_DIR/"reportes",
                               datetime.date(hoy.year,1,1), hoy)
        return jsonify({"ok": True, "msg": f"Reporte guardado: {ruta}", "ruta": ruta})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})

def _generar_excel(datos, carpeta, fecha_ini, fecha_fin):
    carpeta.mkdir(parents=True, exist_ok=True)
    wb  = openpyxl.Workbook()
    fh  = PatternFill("solid", fgColor="1F3864")
    fnt = Font(color="FFFFFF", bold=True)
    emi = [d for d in datos if d.get("tipo")=="Emitida"]
    rec = [d for d in datos if d.get("tipo")=="Recibida"]
    suma = lambda l: round(sum(d.get("total",0) for d in l), 2)
    ws = wb.active; ws.title = "Resumen"
    for r,(k,v) in enumerate([
        ("Periodo",f"{fecha_ini} -> {fecha_fin}"),
        ("Facturas emitidas",len(emi)),("Total emitido MXN",suma(emi)),
        ("Facturas recibidas",len(rec)),("Total recibido MXN",suma(rec)),
        ("Balance",suma(emi)-suma(rec))
    ],1):
        ws.cell(r,1,k).font=Font(bold=True); ws.cell(r,2,v)
    COLS=["tipo","uuid","fecha","rfc_emisor","nombre_emisor","rfc_receptor",
          "nombre_receptor","descripcion","subtotal","total","moneda","tipo_comp","archivo"]
    for nombre,filas in [("Emitidas",emi),("Recibidas",rec)]:
        wsd=wb.create_sheet(nombre)
        for c,col in enumerate(COLS,1):
            cell=wsd.cell(1,c,col.replace("_"," ").title())
            cell.fill=fh; cell.font=fnt
        for r,d in enumerate(filas,2):
            for c,col in enumerate(COLS,1): wsd.cell(r,c,d.get(col,""))
        wsd.auto_filter.ref=wsd.dimensions
    ruta=carpeta/f"CFDIs_{fecha_ini}_{fecha_fin}.xlsx"
    wb.save(ruta); _emit(f"Excel: {ruta}","ok")
    return str(ruta)

@app.route("/api/sistema/abrir-carpeta", methods=["POST"])
def abrir_carpeta():
    try:
        os.startfile(str(DATA_DIR))
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})

@app.route("/api/sistema/log", methods=["GET"])
def get_log():
    try:
        lines = LOG_FILE.read_text("utf-8").splitlines()[-100:] if LOG_FILE.exists() else []
        return jsonify({"ok": True, "lines": lines})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})


# ── Conciliación ──────────────────────────────────────────────
CONCIL_FILE = DATA_DIR / "conciliacion.json"

CATEGORIAS_DEFAULT = [
    {"id": "honorarios", "nombre": "Honorarios profesionales", "color": "purple"},
    {"id": "gastos_med", "nombre": "Gastos médicos",           "color": "teal"},
    {"id": "nomina",     "nombre": "Nómina",                   "color": "green"},
    {"id": "compras",    "nombre": "Compras y mercancías",     "color": "coral"},
    {"id": "sin_cat",    "nombre": "Sin clasificar",           "color": "gray"},
]

def _load_concil():
    if CONCIL_FILE.exists():
        try:
            return json.loads(CONCIL_FILE.read_text("utf-8"))
        except Exception:
            pass
    return {"categorias": CATEGORIAS_DEFAULT, "clasificaciones": {}}

def _save_concil(data):
    CONCIL_FILE.write_text(
        json.dumps(data, indent=2, ensure_ascii=False), "utf-8"
    )

@app.route("/api/conciliacion", methods=["GET"])
def get_conciliacion():
    """Devuelve categorías + facturas con su clasificación."""
    try:
        concil  = _load_concil()
        clases  = concil.get("clasificaciones", {})
        cats    = concil.get("categorias", CATEGORIAS_DEFAULT)

        xmls = list(DATA_DIR.rglob("*.xml"))

        # GASTOS: facturas recibidas (alguien nos cobró)
        recibidas = [_parsear(x, "Recibida") for x in xmls if "recibidas" in x.parts]
        recibidas = [f for f in recibidas if "error" not in f]

        # INGRESOS: facturas emitidas (nosotros cobramos)
        emitidas = [_parsear(x, "Emitida") for x in xmls if "emitidas" in x.parts]
        emitidas = [f for f in emitidas if "error" not in f]

        # Clasificar TODAS las facturas por uso_vista del UsoCFDI
        # Las facturas recibidas con CN01 (nómina) van a INGRESOS
        # Las facturas emitidas siempre van a INGRESOS
        todos = recibidas + emitidas
        gastos_list   = []
        ingresos_list = []

        for f in todos:
            uso_cfdi = f.get("uso_cfdi", "")
            uso_info = USO_CFDI_CAT.get(uso_cfdi, {})
            vista    = f.get("uso_vista", uso_info.get("vista", "gasto"))

            if f.get("tipo") == "Emitida":
                vista = "ingreso"   # Emitidas siempre son ingresos

            f["categoria"]  = uso_cfdi or "sin_cat"
            f["cat_desc"]   = uso_info.get("desc", "Sin clasificar")
            f["cat_origen"] = "auto"
            f["vista"]      = vista

            if vista == "ingreso":
                ingresos_list.append(f)
            else:
                gastos_list.append(f)

        # Agrupar gastos por UsoCFDI para el sidebar
        grupos_gasto = {}
        for f in gastos_list:
            cid = f.get("uso_cfdi") or "sin_cat"
            if cid not in grupos_gasto:
                grupos_gasto[cid] = {
                    "id":    cid,
                    "desc":  USO_CFDI_CAT.get(cid, {}).get("desc", "Sin clasificar"),
                    "count": 0,
                    "monto": 0.0,
                }
            grupos_gasto[cid]["count"] += 1
            grupos_gasto[cid]["monto"] = round(grupos_gasto[cid]["monto"] + f.get("total", 0), 2)

        # Agrupar ingresos por UsoCFDI
        grupos_ingreso = {}
        for f in ingresos_list:
            cid = f.get("uso_cfdi") or "sin_cat"
            if cid not in grupos_ingreso:
                grupos_ingreso[cid] = {
                    "id":    cid,
                    "desc":  USO_CFDI_CAT.get(cid, {}).get("desc", "Sin clasificar"),
                    "count": 0,
                    "monto": 0.0,
                }
            grupos_ingreso[cid]["count"] += 1
            grupos_ingreso[cid]["monto"] = round(grupos_ingreso[cid]["monto"] + f.get("total", 0), 2)

        # Totales de impuestos (sumados de todas las facturas)
        impuestos = {
            "ret_isr":         round(sum(f.get("ret_isr",0)        for f in gastos_list), 2),
            "ret_iva":         round(sum(f.get("ret_iva",0)        for f in gastos_list), 2),
            "iva_trasladado":  round(sum(f.get("iva_trasladado",0) for f in gastos_list), 2),
            "ret_ieps":        round(sum(f.get("ret_ieps",0)       for f in gastos_list), 2),
        }

        # Facturas canceladas: TipoDeComprobante = E (egreso/nota de crédito)
        # La API del SAT no descarga canceladas directamente, pero podemos
        # identificar CFDIs de egreso que referencian otros (nota de crédito)
        canceladas = [f for f in todos if f.get("tipo_comprobante") == "E"]

        # Totales de gastos por categoría
        totales = {}
        for cat in cats:
            cid   = cat["id"]
            grupo = [f for f in recibidas if f.get("categoria") == cid]
            totales[cid] = {
                "count": len(grupo),
                "monto": round(sum(f.get("total", 0) for f in grupo), 2),
            }

        total_gastos   = round(sum(f.get("total", 0) for f in gastos_list),   2)
        total_ingresos = round(sum(f.get("total", 0) for f in ingresos_list), 2)

        return jsonify({
            "ok":               True,
            "facturas":         gastos_list,
            "ingresos":         ingresos_list,
            "grupos_gasto":     list(grupos_gasto.values()),
            "grupos_ingreso":   list(grupos_ingreso.values()),
            "impuestos":        impuestos,
            "canceladas":       canceladas,
            "total_gastos":     total_gastos,
            "total_ingresos":   total_ingresos,
            "balance":          round(total_ingresos - total_gastos, 2),
            "total_count":      len(gastos_list),
            "ingresos_count":   len(ingresos_list),
            "canceladas_count": len(canceladas),
        })
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})

@app.route("/api/conciliacion/clasificar", methods=["POST"])
def clasificar_factura():
    """Asigna una categoría a una factura por UUID."""
    try:
        d    = request.get_json() or {}
        uuid = d.get("uuid", "").strip()
        cat  = d.get("categoria", "sin_cat").strip()
        if not uuid:
            return jsonify({"ok": False, "msg": "UUID requerido."})
        concil = _load_concil()
        concil["clasificaciones"][uuid] = cat
        _save_concil(concil)
        return jsonify({"ok": True, "uuid": uuid, "categoria": cat})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})

@app.route("/api/conciliacion/clasificar-lote", methods=["POST"])
def clasificar_lote():
    """Clasifica múltiples facturas a la vez: [{"uuid": ..., "categoria": ...}]"""
    try:
        items  = request.get_json() or []
        concil = _load_concil()
        n = 0
        for item in items:
            uuid = item.get("uuid", "").strip()
            cat  = item.get("categoria", "sin_cat").strip()
            if uuid:
                concil["clasificaciones"][uuid] = cat
                n += 1
        _save_concil(concil)
        return jsonify({"ok": True, "clasificadas": n})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})

@app.route("/api/conciliacion/categorias", methods=["GET"])
def get_categorias():
    concil = _load_concil()
    return jsonify({"ok": True, "categorias": concil.get("categorias", CATEGORIAS_DEFAULT)})

@app.route("/api/conciliacion/categorias", methods=["POST"])
def save_categorias():
    """Guarda la lista completa de categorías (crear / renombrar / reordenar)."""
    try:
        cats   = request.get_json() or []
        concil = _load_concil()
        # Asegurar que siempre exista sin_cat
        ids = [c["id"] for c in cats]
        if "sin_cat" not in ids:
            cats.append({"id": "sin_cat", "nombre": "Sin clasificar", "color": "gray"})
        concil["categorias"] = cats
        _save_concil(concil)
        return jsonify({"ok": True, "categorias": cats})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})

@app.route("/api/conciliacion/exportar-excel", methods=["POST"])
def exportar_conciliacion_excel():
    """Genera Excel con una hoja por categoría."""
    try:
        concil    = _load_concil()
        clases    = concil.get("clasificaciones", {})
        cats      = concil.get("categorias", CATEGORIAS_DEFAULT)
        xmls      = list(DATA_DIR.rglob("*.xml"))
        recibidas = [_parsear(x, "Recibida") for x in xmls if "recibidas" in x.parts]
        recibidas = [f for f in recibidas if "error" not in f]
        for f in recibidas:
            f["categoria"] = clases.get(f.get("uuid",""), "sin_cat")

        wb  = openpyxl.Workbook()
        fh  = PatternFill("solid", fgColor="1F3864")
        fnt = Font(color="FFFFFF", bold=True)
        COLS = ["uuid","fecha","rfc_emisor","nombre_emisor",
                "descripcion","subtotal","total","moneda","categoria"]

        ws = wb.active; ws.title = "Resumen"
        ws.cell(1,1,"Categoría").font = Font(bold=True)
        ws.cell(1,2,"Facturas").font  = Font(bold=True)
        ws.cell(1,3,"Total MXN").font = Font(bold=True)
        for r, cat in enumerate(cats, 2):
            grupo = [f for f in recibidas if f.get("categoria") == cat["id"]]
            ws.cell(r,1,cat["nombre"])
            ws.cell(r,2,len(grupo))
            ws.cell(r,3,round(sum(f.get("total",0) for f in grupo), 2))

        for cat in cats:
            grupo = [f for f in recibidas if f.get("categoria") == cat["id"]]
            if not grupo:
                continue
            nombre_hoja = cat["nombre"][:31]
            wsd = wb.create_sheet(nombre_hoja)
            for c, col in enumerate(COLS, 1):
                cell = wsd.cell(1, c, col.replace("_"," ").title())
                cell.fill = fh; cell.font = fnt
            for r, f in enumerate(grupo, 2):
                for c, col in enumerate(COLS, 1):
                    wsd.cell(r, c, f.get(col, ""))
            wsd.auto_filter.ref = wsd.dimensions

        hoy  = datetime.date.today()
        ruta = DATA_DIR / "reportes" / f"Conciliacion_{hoy}.xlsx"
        ruta.parent.mkdir(parents=True, exist_ok=True)
        wb.save(ruta)
        return jsonify({"ok": True, "msg": f"Excel guardado: {ruta}", "ruta": str(ruta)})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})


@app.route("/api/diagnostico/xml", methods=["GET"])
def diagnostico_xml():
    """Muestra los primeros 5 XMLs parseados para verificar que los campos se lean bien."""
    xmls      = list(DATA_DIR.rglob("*.xml"))[:10]
    resultado = []
    for x in xmls:
        p = _parsear(x, "rec" if "recibidas" in x.parts else "emi")
        resultado.append({
            "archivo":       p.get("archivo"),
            "error":         p.get("error"),
            "uuid":          p.get("uuid", "")[:12],
            "nombre_emisor": p.get("nombre_emisor"),
            "rfc_emisor":    p.get("rfc_emisor"),
            "uso_cfdi":      p.get("uso_cfdi"),
            "uso_cfdi_desc": p.get("uso_cfdi_desc"),
            "total":         p.get("total"),
            "tipo_comp":     p.get("tipo_comprobante"),
        })
    return jsonify({"total_xmls": len(list(DATA_DIR.rglob("*.xml"))), "muestra": resultado})


@app.route("/api/conciliacion/xml/<uuid>", methods=["GET"])
def descargar_xml(uuid):
    """Descarga el archivo XML de un CFDI por UUID."""
    for xml in DATA_DIR.rglob("*.xml"):
        contenido = xml.read_text("utf-8", errors="replace")
        if uuid.lower() in contenido.lower():
            return send_file(
                str(xml),
                mimetype="application/xml",
                as_attachment=True,
                download_name=f"{uuid[:8]}.xml"
            )
    return jsonify({"ok": False, "msg": "XML no encontrado."}), 404

@app.route("/api/impuestos", methods=["GET"])
def get_impuestos():
    """Desglose de impuestos de todas las facturas recibidas."""
    try:
        xmls      = list(DATA_DIR.rglob("*.xml"))
        recibidas = [_parsear(x, "Recibida") for x in xmls if "recibidas" in x.parts]
        recibidas = [f for f in recibidas if "error" not in f]

        ret_isr  = round(sum(f.get("ret_isr",0)        for f in recibidas), 2)
        ret_iva  = round(sum(f.get("ret_iva",0)        for f in recibidas), 2)
        iva_tra  = round(sum(f.get("iva_trasladado",0) for f in recibidas), 2)
        ret_ieps = round(sum(f.get("ret_ieps",0)       for f in recibidas), 2)

        # Desglose por factura con impuestos
        con_impuestos = [
            {
                "uuid":           f.get("uuid","")[:16],
                "nombre_emisor":  f.get("nombre_emisor",""),
                "fecha":          f.get("fecha",""),
                "total":          f.get("total",0),
                "iva_trasladado": f.get("iva_trasladado",0),
                "ret_isr":        f.get("ret_isr",0),
                "ret_iva":        f.get("ret_iva",0),
                "ret_ieps":       f.get("ret_ieps",0),
                "uso_cfdi":       f.get("uso_cfdi",""),
            }
            for f in recibidas
            if any([f.get("ret_isr"), f.get("ret_iva"),
                    f.get("iva_trasladado"), f.get("ret_ieps")])
        ]

        return jsonify({
            "ok":          True,
            "resumen":     {
                "ret_isr":          ret_isr,
                "ret_iva":          ret_iva,
                "iva_trasladado":   iva_tra,
                "ret_ieps":         ret_ieps,
                "total_retenciones":round(ret_isr + ret_iva + ret_ieps, 2),
            },
            "detalle":     con_impuestos,
            "total_facturas": len(recibidas),
        })
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})

# ── Arranque ──────────────────────────────────────────────────
if __name__ == "__main__":
    if not GUI_FILE.exists():
        print(f"[ERROR] GUI no encontrada: {GUI_FILE}")
        input("Presiona Enter para cerrar...")
        sys.exit(1)

    url = f"http://localhost:{PORT}"
    print(f"\n  ContaSAT v3.0 iniciando en {url}")
    print(f"  Abriendo en el navegador...")
    print(f"  Para cerrar la aplicacion cierra esta ventana.\n")

    threading.Timer(1.5, lambda: webbrowser.open(url)).start()
    app.run(host="127.0.0.1", port=PORT, debug=False, use_reloader=False)
